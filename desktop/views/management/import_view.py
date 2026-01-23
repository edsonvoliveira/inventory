# desktop/views/management/import_view.py

"""
Responsibilities:
- Render a generic import flow for CSV/XLSX.
- Support column mapping and validation for products.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import flet as ft

from desktop.core.strings import IMPORT_TITLE
from desktop.data.db.connection import get_connection
from desktop.data.repositories.products_repo import ProductsRepo


REQUIRED_FIELDS = [
    ("sku", "SKU"),
    ("name", "Nome"),
]

OPTIONAL_FIELDS = [
    ("uom_base", "UOM Base"),
    ("uom_inventory", "UOM Inventário"),
    ("description", "Descrição"),
    ("conversion_factor", "Fator Conversão"),
    ("cost_price", "Custo"),
    ("is_sensitive", "Sensível"),
    ("serial_number_enabled", "Serial"),
    ("is_active", "Ativo"),
]


@dataclass
class ImportState:
    step: int = 0
    file_path: str | None = None
    columns: list[str] = None
    rows: list[dict[str, Any]] = None
    mapping: dict[str, str] = None
    validation_errors: list[str] = None
    import_errors: list[str] = None
    ignore_existing_skus: bool = False
    existing_skus: set[str] = None
    skipped_existing: int = 0
    imported_count: int = 0
    report_path: str | None = None

    def __post_init__(self) -> None:
        self.columns = []
        self.rows = []
        self.mapping = {}
        self.validation_errors = []
        self.import_errors = []
        self.ignore_existing_skus = False
        self.existing_skus = set()
        self.skipped_existing = 0
        self.imported_count = 0
        self.report_path = None


def _snack(page: ft.Page, message: str) -> None:
    page.snack_bar = ft.SnackBar(content=ft.Text(message), open=True, duration=2000)
    page.update()


def _parse_csv(path: str) -> tuple[list[str], list[dict[str, Any]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=dialect)
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames or []), rows


def _parse_xlsx(path: str) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on optional lib
        raise RuntimeError("openpyxl não instalado para XLSX.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
    rows = []
    for row in rows_iter:
        data = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        rows.append(data)
    return headers, rows


def _parse_file(path: str) -> tuple[list[str], list[dict[str, Any]]]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return _parse_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path)
    raise RuntimeError("Formato não suportado. Use CSV ou XLSX.")


def _auto_map(columns: list[str]) -> dict[str, str]:
    normalized = {c.lower().strip(): c for c in columns}
    mapping: dict[str, str] = {}
    for field, _label in REQUIRED_FIELDS + OPTIONAL_FIELDS:
        if field in normalized:
            mapping[field] = normalized[field]
    return mapping


def _parse_bool(value: Any) -> int:
    if value is None:
        return 0
    text = str(value).strip().lower()
    return 1 if text in {"1", "true", "yes", "sim", "y", "s"} else 0


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def render_import_view(page: ft.Page, on_refresh):
    state = ImportState()

    header = ft.Text(IMPORT_TITLE, size=28, weight=ft.FontWeight.BOLD)
    step_title = ft.Text("1 Selecionar arquivo", size=22, weight=ft.FontWeight.BOLD)

    file_info = ft.Text("Nenhum arquivo selecionado.")
    preview_table = ft.DataTable(columns=[ft.DataColumn(ft.Text("Preview"))], rows=[])

    file_picker = ft.FilePicker()
    page.overlay.append(file_picker)

    def _set_step(value: int) -> None:
        state.step = value
        step_title.value = [
            "1 - Selecionar arquivo",
            "2 - Mapear colunas",
            "3 - Validar dados",
            "4 - Importar",
        ][state.step]
        step1.visible = state.step == 0
        step2.visible = state.step == 1
        step3.visible = state.step == 2
        step4.visible = state.step == 3
        page.update()

    def _build_preview():
        preview_table.columns = [ft.DataColumn(ft.Text(c)) for c in state.columns[:5]]
        preview_rows = state.rows[:5]
        preview_table.rows = [
            ft.DataRow(cells=[ft.DataCell(ft.Text(str(r.get(c, "")))) for c in state.columns[:5]])
            for r in preview_rows
        ]

    def _select_file(_e):
        file_picker.pick_files(
            allow_multiple=False,
            allowed_extensions=["csv", "xlsx", "xlsm"],
        )

    def _on_file_result(e: ft.FilePickerResultEvent):
        if not e.files:
            return
        state.file_path = e.files[0].path
        try:
            state.columns, state.rows = _parse_file(state.file_path)
        except Exception as exc:
            _snack(page, f"Erro ao ler arquivo: {exc}")
            return
        state.mapping = _auto_map(state.columns)
        file_info.value = f"Arquivo: {Path(state.file_path).name} | Linhas: {len(state.rows)}"
        _build_preview()
        _snack(page, "Arquivo carregado.")
        page.update()

    file_picker.on_result = _on_file_result

    def _next_from_file():
        if not state.file_path:
            _snack(page, "Selecione um arquivo antes de avançar.")
            return
        mapping_view.content = _build_mapping_controls()
        _set_step(1)

    mapping_controls: dict[str, ft.Dropdown] = {}
    mapping_view = ft.Container()

    def _build_mapping_controls():
        mapping_controls.clear()
        required_set = {field for field, _label in REQUIRED_FIELDS}
        system_options = [
            ft.dropdown.Option("__ignore__", "Ignorar coluna"),
            *[ft.dropdown.Option(field, label) for field, label in (REQUIRED_FIELDS + OPTIONAL_FIELDS)],
        ]
        auto_map = _auto_map(state.columns)
        selected_by_file: dict[str, str] = {}
        for field, col in auto_map.items():
            if col:
                selected_by_file[col] = field

        row_height = 44
        left_controls = []
        right_controls = []
        for col in state.columns:
            left_controls.append(
                ft.Container(
                    content=ft.TextField(
                        value=col,
                        read_only=True,
                        width=260,
                        height=row_height,
                    ),
                    height=row_height,
                    alignment=ft.alignment.center_left,
                )
            )
            dd = ft.Dropdown(
                options=system_options,
                width=260,
            )
            dd.value = selected_by_file.get(col, "__ignore__")
            mapping_controls[col] = dd
            right_controls.append(
                ft.Container(
                    content=dd,
                    height=row_height,
                    alignment=ft.alignment.center_left,
                )
            )

        return ft.Row(
            [
                ft.Column(
                    [
                        ft.Text("Coluna do arquivo", weight=ft.FontWeight.BOLD),
                        *left_controls,
                    ],
                    spacing=6,
                ),
                ft.Column(
                    [
                        ft.Text("Campos do Sistema", weight=ft.FontWeight.BOLD),
                        *right_controls,
                    ],
                    spacing=6,
                ),
            ],
            spacing=24,
            vertical_alignment=ft.CrossAxisAlignment.START,
        )

    def _next_from_mapping():
        state.mapping = {}
        used_fields: set[str] = set()
        for col, dropdown in mapping_controls.items():
            value = dropdown.value or "__ignore__"
            if value == "__ignore__":
                continue
            if value in used_fields:
                _snack(page, f"Campo duplicado no mapeamento: {value}")
                return
            used_fields.add(value)
            state.mapping[value] = col

        for field, _label in REQUIRED_FIELDS:
            if field not in state.mapping:
                _snack(page, f"Mapeie o campo obrigatório: {field}")
                return
        _set_step(2)
        _validate_data()

    validation_summary = ft.Text("")
    validation_list = ft.Column([])
    ignore_existing_checkbox = ft.Checkbox(
        label="Ignorar SKUs já existentes e importar os demais",
        value=False,
    )

    def _validate_data():
        errors: list[str] = []
        sku_map = state.mapping.get("sku")
        name_map = state.mapping.get("name")

        existing_skus: set[str] = set()
        conn = get_connection()
        try:
            rows = conn.execute("SELECT sku FROM products_local").fetchall()
            existing_skus = {r[0] for r in rows if r and r[0]}
        finally:
            conn.close()
        state.existing_skus = existing_skus

        seen_skus: set[str] = set()
        for idx, row in enumerate(state.rows, start=2):
            sku = str(row.get(sku_map, "")).strip() if sku_map else ""
            name = str(row.get(name_map, "")).strip() if name_map else ""
            if not sku:
                errors.append(f"Linha {idx}: SKU vazio.")
            if not name:
                errors.append(f"Linha {idx}: Nome vazio.")
            if sku:
                if sku in seen_skus:
                    errors.append(f"Linha {idx}: SKU duplicado no arquivo ({sku}).")
                if sku in existing_skus:
                    errors.append(f"Linha {idx}: SKU já existe no local ({sku}).")
                seen_skus.add(sku)

        state.validation_errors = errors
        if errors:
            validation_summary.value = f"Encontrados {len(errors)} erros."
            validation_list.controls = [ft.Text(err, size=12) for err in errors[:20]]
        else:
            validation_summary.value = "Nenhum erro encontrado. Pronto para importar."
            validation_list.controls = [ft.Text("OK", size=12)]
        page.update()

    def _next_from_validation():
        state.ignore_existing_skus = bool(ignore_existing_checkbox.value)
        if state.validation_errors:
            if state.ignore_existing_skus:
                remaining = [
                    err for err in state.validation_errors
                    if "SKU já existe no local" not in err
                ]
                if remaining:
                    _snack(page, "Corrija os erros antes de importar.")
                    return
            else:
                _snack(page, "Corrija os erros antes de importar.")
                return
        _set_step(3)

    import_summary = ft.Text("")
    import_details = ft.Column([])
    report_button = ft.ElevatedButton("Salvar relatório de erros")
    required_hint = ft.Text(
        "Obrigatórios: SKU, Nome",
        size=12,
        color=ft.Colors.ORANGE_700,
    )
    defaults_hint = ft.Text(
        "Sem UOM base/inventario/fator, usa: UN, UN e 1.",
        size=12,
        color=ft.Colors.GREY_600,
    )

    def _build_report_text() -> str:
        conn = get_connection()
        try:
            company_name = "n/a"
            user_email = "n/a"
            company_id = get_meta("company_id", conn)
            user_id = get_meta("user_server_id", conn)
            if company_id:
                row = conn.execute(
                    "SELECT name FROM companies_local WHERE server_id = ?",
                    (company_id,),
                ).fetchone()
                if row:
                    company_name = row[0]
            if user_id:
                row = conn.execute(
                    "SELECT email FROM users_local WHERE server_id = ?",
                    (user_id,),
                ).fetchone()
                if row:
                    user_email = row[0]
        finally:
            conn.close()

        header_lines = [
            f"Data: {datetime.now(timezone.utc).isoformat()}",
            f"Empresa: {company_name}",
            f"Utilizador: {user_email}",
            f"Importados: {state.imported_count}",
            f"Ignorados (SKU existente): {state.skipped_existing}",
            f"Erros: {len(state.import_errors)}",
            "",
        ]
        return "\n".join(header_lines + state.import_errors)

    def _do_import(_e):
        errors = []
        skipped_existing = 0
        imported_count = 0
        conn = get_connection()
        try:
            repo = ProductsRepo(conn)
            for idx, row in enumerate(state.rows, start=2):
                data: dict[str, Any] = {}
                for field, _label in REQUIRED_FIELDS + OPTIONAL_FIELDS:
                    col = state.mapping.get(field)
                    if not col:
                        continue
                    value = row.get(col)
                    if field in {"conversion_factor", "cost_price"}:
                        data[field] = _parse_float(value)
                    elif field in {"is_sensitive", "serial_number_enabled", "is_active"}:
                        data[field] = _parse_bool(value)
                    else:
                        data[field] = str(value).strip() if value is not None else ""
                if not data.get("uom_base"):
                    data["uom_base"] = "UN"
                if not data.get("uom_inventory"):
                    data["uom_inventory"] = data["uom_base"]
                if data.get("conversion_factor") in (None, ""):
                    data["conversion_factor"] = 1
                if state.ignore_existing_skus:
                    sku_value = data.get("sku")
                    if sku_value and sku_value in state.existing_skus:
                        skipped_existing += 1
                        continue
                try:
                    repo.create(data)
                    imported_count += 1
                except Exception as exc:
                    errors.append(f"Linha {idx}: {exc}")
            conn.commit()
        finally:
            conn.close()
        state.import_errors = errors
        state.skipped_existing = skipped_existing
        state.imported_count = imported_count
        if errors:
            import_summary.value = (
                f"Importação finalizada com {len(errors)} erros. "
                f"Importados: {imported_count}. Ignorados: {skipped_existing}."
            )
            import_details.controls = [ft.Text(err, size=12) for err in errors[:20]]
            report_button.visible = True
        else:
            import_summary.value = (
                f"Importação concluída com sucesso. "
                f"Importados: {imported_count}. Ignorados: {skipped_existing}."
            )
            import_details.controls = [ft.Text("OK", size=12)]
            report_button.visible = False
        _snack(page, "Importação finalizada.")
        page.update()

    def _save_report(_e):
        if not state.import_errors:
            _snack(page, "Nao ha erros para salvar.")
            return
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        save_report_picker.save_file(
            dialog_title="Salvar relatório de erros",
            file_name=f"import_errors_{timestamp}.txt",
        )

    step1 = ft.Column(
        [
            ft.ElevatedButton("Selecionar arquivo", on_click=_select_file),
            file_info,
            preview_table,
            ft.ElevatedButton("Avançar", on_click=lambda e: _next_from_file()),
        ],
        spacing=12,
        visible=True,
    )

    step2 = ft.Column(
        [
            ft.Text("2 - Mapear colunas", weight=ft.FontWeight.BOLD),
            ft.Text("Mapeie os campos obrigatórios para Produtos."),
            mapping_view,
            required_hint,
            defaults_hint,
            ft.ElevatedButton("Avançar", on_click=lambda e: _next_from_mapping()),
        ],
        spacing=12,
        visible=False,
        scroll=ft.ScrollMode.AUTO,
    )

    step3 = ft.Column(
        [
            validation_summary,
            validation_list,
            ignore_existing_checkbox,
            ft.ElevatedButton("Avançar", on_click=lambda e: _next_from_validation()),
        ],
        spacing=12,
        visible=False,
        scroll=ft.ScrollMode.AUTO,
    )

    step4 = ft.Column(
        [
            ft.Text("A importação criará produtos locais e enviará via sync."),
            ft.ElevatedButton("Importar", on_click=_do_import),
            import_summary,
            import_details,
            report_button,
        ],
        spacing=12,
        visible=False,
        scroll=ft.ScrollMode.AUTO,
    )

    report_button.on_click = _save_report
    report_button.visible = False

    save_report_picker = ft.FilePicker()

    def _on_report_saved(e: ft.FilePickerResultEvent):
        if not e.path:
            return
        report_path = Path(e.path)
        report_path.write_text(_build_report_text(), encoding="utf-8")
        state.report_path = str(report_path)
        _snack(page, f"Relatorio salvo: {report_path}")

    save_report_picker.on_result = _on_report_saved
    page.overlay.append(save_report_picker)

    _set_step(0)

    return ft.Column(
        [
            header,
            step_title,
            step1,
            step2,
            step3,
            step4,
        ],
        spacing=16,
        expand=True,
        scroll=ft.ScrollMode.AUTO,
    )
